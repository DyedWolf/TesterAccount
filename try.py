from openpyxl import *


class execl_op:
    # 绑定
    def __init__(self, fname, sname):
        self.fname = fname
        self.wb = load_workbook(self.fname)
        print(self.wb)
        self.ws = self.wb[sname]

    def del_flsh_rows_cols(self, cold):  # 基于文件名和表格名删除一些行和一些列
        """基于文件名和表格名删除一些行和一些列
        要删的行序数放在rowd表格中，要删的列序数放在cold表格中
        本程序的关键是删除的行或列序数都必须是从大的开始删除，这样才不会乱序"""
        # rowd = sorted(rowd, reverse=True)
        cold = sorted(cold, reverse=True)
        # for r in rowd:
        #     ws.delete_rows(r)
        for c in cold:
            self.ws.delete_cols(c)
        self.wb.save(self.fname)  # 记得要保存。

    # 获取某列所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data



    def create_sheet(self):

        # 创建一个excel工作簿
        wb = Workbook()

        # 可以通过参数来指定sheet位置,0表示第一个
        first_sheet = wb.create_sheet("Android分支测试进度", 0)

        for i in range(1, 2):
            first_sheet.append(execl_op.get_row_value(1))

        # sheet可以通过append来添加一行(row)
        for i in range(2, execl_op.get_col_value(2).index('iOS')+1):
            first_sheet.append(execl_op.get_row_value(i))

        first_sheet = wb.create_sheet("iOS分支测试进度", 1)

        for i in range(1, 2):
            first_sheet.append(execl_op.get_row_value(1))

        for i in range(execl_op.get_col_value(2).index('iOS') , execl_op.get_col_value(2).index(None)+1):
            first_sheet.append(execl_op.get_row_value(i))


        wb.save('/TesterAccount/AccountManage/static/excel/schedule.xlsx')
        wb.close()


if __name__ == '__main__':
    fname = "/TesterAccount/AccountManage/static/excel/2023年互动项目资源排期表.xlsx"
    sname = "YO语音1.14"
    execl_op = execl_op(fname, sname)
    execl_op.del_flsh_rows_cols([2, 5, 6, 7, 8, 9, 10, 11, 12, 13, 16, 17, 18, 19, 26, 27, 28, 29, 30])
    second_col = execl_op.get_col_value(2)
    #
    # end_row = int(second_col.index('iOS'))
    # exce_name = "Android分支测试进度"
    # execl_op.create_sheet(exce_name, end_row, 1)
    #
    # begin_row = int(second_col.index('iOS'))
    # exce_name = "iOS分支测试进度"
    # end_row = int((second_col.index(None)))
    execl_op.create_sheet()
