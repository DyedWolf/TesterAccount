# from openpyxl import load_workbook
#
# # 打开excel表格
# data_excel = load_workbook('/TesterAccount/AccountManage/static/excel/2023年互动项目资源排期表.xlsx')
#
# wb = load_workbook('/TesterAccount/AccountManage/static/excel/2023年互动项目资源排期表.xlsx', read_only=True, data_only=True)
# ws = wb['YO语音1.14']
# for row in ws.rows:  # 获取每一行的数据
#     print(row)
#     for data in row:  # 获取每一行中单元格的数据
#         print(data.value)  # 打印单元格的值

from openpyxl import load_workbook


# 基于ws删除一些行和一些列，注意没有备份,
# 并且最后没有保存，因为可能ws还有后续操作
# 必须记得最后还要保存。
def del_ws_rows_cols(ws, rowd, cold):  # 删除一些行和一些列，此程序不含保存操作。
    """基于ws删除一些行和一些列
    要删的行序数放在rowd表格中，要删的列序数放在cold表格中
    本程序的关键是删除的行或列序数都必须是从大的开始删除，这样才不会乱序"""
    # wb = load_workbook(flname)
    # ws = wb[sheetname]
    rowd = sorted(rowd, reverse=True)  # 确保大的行数首先删除
    cold = sorted(cold, reverse=True)  # 确保大的列数首先删除
    for r in rowd:  # rowd格式如：[1,3,5],表示要删除第1、3、5共三行。
        ws.delete_rows(r)
    for c in cold:  # cold格式如：[2,6,10],表示要删除第2、6、10共三列
        ws.delete_cols(c)
    # wb.save(flname)  # 记得要保存。但此程序中不出现保存操作。
    # 因为在打开ws的情况下可能还要进行其他操作，然后再一次性进行保存。


# 基于文件名和表格名删除一些行和一些列，注意没有备份。
# flsh是指文件名flname和表格名sheetname
def del_flsh_rows_cols(flname, sheetname, cold):  # 基于文件名和表格名删除一些行和一些列
    """基于文件名和表格名删除一些行和一些列
    要删的行序数放在rowd表格中，要删的列序数放在cold表格中
    本程序的关键是删除的行或列序数都必须是从大的开始删除，这样才不会乱序"""
    wb = load_workbook(flname)
    ws = wb[sheetname]
    # rowd = sorted(rowd, reverse=True)
    cold = sorted(cold, reverse=True)
    # for r in rowd:
    #     ws.delete_rows(r)
    for c in cold:
        ws.delete_cols(c)
    wb.save(flname)  # 记得要保存。


if __name__ == "__main__":
    flname = '/TesterAccount/AccountManage/static/excel/2023年互动项目资源排期表.xlsx'
    sheetname = 'YO语音1.14'
    del_flsh_rows_cols(flname, sheetname, [2, 5, 6, 7, 8, 9, 10, 11, 12, 13, 16, 17, 18, 19, 26, 27, 28, 29, 30])


import openpyxl


class execl_op:
    # 绑定
    def __init__(self, fname, sname):
        self.fname = fname
        self.wb = openpyxl.load_workbook(self.fname)
        self.ws = self.wb[sname]
        # sheets = self.wb.sheetnames
        # self.sheet = sheets[0]
        # self.ws = self.wb[self.sheet]

    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取表格某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 获取某列所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 根据单元格的值获取列号
    def get_cell_column(self):
        column = self.ws.max_column
        for i in range(1, column + 1):
            cell_value = self.ws.cell(row=1, column=i).value
            if cell_value == 'SN' or cell_value == 'sn':
                return i

    # 设置某个单元格的值
    def set_cell_value(self, row, colum, cellvalue):
        def set_cell_value(self, row, colunm, cellvalue):
            try:
                self.ws.cell(row=row, column=colunm).value = cellvalue
                self.wb.save(self.file)
            except:
                self.ws.cell(row=row, column=colunm).value = "writefail"
                self.wb.save(self.file)


if __name__ == '__main__':
    # num = input("请输入多少台设备：")
    fname = "/TesterAccount/AccountManage/static/excel/2023年互动项目资源排期表.xlsx"
    sname = "YO语音1.14"
    # first_row = execl_op(fname).get_row_value(1)
    # print(first_row)
    second_col = execl_op(fname, sname).get_col_value(2)
    print(second_col)
    # column_num = execl_op(fname).get_cell_column()
    # print('SN的列数是：%s' % column_num)
    for i, value in enumerate(second_col):
        if not value:
            # print(i)
            second_col.remove(second_col[i])
    print(second_col.index(None))
    print(second_col.index('iOS'))
    print(len(second_col))
