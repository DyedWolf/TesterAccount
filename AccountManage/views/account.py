import os
from openpyxl import load_workbook

from django.db import connection
from django.shortcuts import render, HttpResponse, redirect
from django.http.response import JsonResponse, FileResponse
from django.views.decorators.csrf import csrf_exempt

from AccountManage import models
from AccountManage.utils.form import Account, Upload
from AccountManage.utils.pagination import Pagination


def account_list(request):
    """ 用户列表 """
    if request.method == "GET":
        form_list = Account()
        form = Upload()
        data_dict = {}
        aid = request.session["info"]["id"]
        # print(aid)
        txt_search = request.GET.get("txtSearch", "")
        search_id = request.GET.get("idSearch", "")
        print(txt_search)
        print(search_id)
        phone_data = models.HDID.objects.filter().values("id", "phone")
        print("phone_data:{}".format(phone_data))

        with connection.cursor() as cursor:
            if search_id == 'YY' and txt_search:
                sql = "SELECT a.id,a.account_uid,a.account_YY,a.account_card,a.account_pwd,b.tester_name,a.account_remarks FROM (SELECT CASE WHEN iu.`account_belong_id` = '{}' THEN 0 ELSE iu.`account_belong_id` END AS iid,iu.* FROM accountmanage_account iu) a ,accountmanage_testerinfo b WHERE a.`account_belong_id`=b.id AND a.account_YY like '%{}%' ORDER BY a.iid ASC;".format(
                    aid, txt_search)
            elif search_id == 'UID' and txt_search:
                sql = "SELECT a.id,a.account_uid,a.account_YY,a.account_card,a.account_pwd,b.tester_name,a.account_remarks FROM (SELECT CASE WHEN iu.`account_belong_id` = '{}' THEN 0 ELSE iu.`account_belong_id` END AS iid,iu.* FROM accountmanage_account iu) a ,accountmanage_testerinfo b WHERE a.`account_belong_id`=b.id AND a.account_UID like '%{}%' ORDER BY a.iid ASC;".format(
                    aid, txt_search)
            elif search_id == '通行证' and txt_search:
                sql = "SELECT a.id,a.account_uid,a.account_YY,a.account_card,a.account_pwd,b.tester_name,a.account_remarks FROM (SELECT CASE WHEN iu.`account_belong_id` = '{}' THEN 0 ELSE iu.`account_belong_id` END AS iid,iu.* FROM accountmanage_account iu) a ,accountmanage_testerinfo b WHERE a.`account_belong_id`=b.id AND a.account_card like '%{}%' ORDER BY a.iid ASC;".format(
                    aid, txt_search)
            else:
                sql = "SELECT a.id,a.account_uid,a.account_YY,a.account_card,a.account_pwd,b.tester_name,a.account_remarks FROM (SELECT CASE WHEN iu.`account_belong_id` = '{}' THEN 0 ELSE iu.`account_belong_id` END AS iid,iu.* FROM accountmanage_account iu) a ,accountmanage_testerinfo b WHERE a.`account_belong_id`=b.id ORDER BY a.iid ASC;".format(
                    aid)

            # print(sql)
            cursor.execute(sql)
            sql_data = cursor.fetchall()  # 获取一条数据, 使用fetchone()
            colums = [col[0] for col in cursor.description]
            sql_data = [dict(zip(colums, row)) for row in sql_data]
            # print(sql_data)

        # queryset = models.Account.objects.filter(**data_dict).order_by("id")
        page_object = Pagination(request, sql_data)
        page_queryset = page_object.page_queryset
        page_string = page_object.html()
        context = {
            "phone_data": phone_data,
            "form": form,
            "form_list": form_list,
            "form_data": page_queryset,
            "txt_search": txt_search,
            "page_string": page_string,
            "search_id": search_id
        }
        # print(context["form_list"])
        # return render(request, "account_list.html", context)

        return render(request, "account_newList.html", context)


@csrf_exempt
def account_multi(request):
    """ 批量上传 """
    if request.method == "GET":
        form = Upload()
        return render(request, "account_upload.html", {"form": form})
    account_belong_id = request.POST.get('account_belong')
    print(account_belong_id)
    file_object = request.FILES.get("exc")
    print(file_object)
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        data_dict = {
            "account_card": row[0].value,
            "account_pwd": row[1].value,
            "account_uid": row[2].value,
            "account_YY": row[3].value,
            "account_belong_id": account_belong_id
        }
        print(data_dict)
        if None in data_dict.values():
            result = {"status": False, "error": "excel cannot have null value"}
            return JsonResponse(result)
        exits = models.Account.objects.filter(**data_dict).exists()
        if not exits:
            models.Account.objects.create(**data_dict)
    return redirect("/account/list/")


@csrf_exempt
def account_delete(request):
    """删除测试账号"""
    aid = request.GET.get("aid")
    print(aid)
    models.Account.objects.filter(id=aid).delete()
    result = {"status": True}
    return JsonResponse(result)


@csrf_exempt
def account_add(request):
    """添加测试账号"""
    form = Account(data=request.POST)
    print(request.POST)
    if form.is_valid():
        # form.save()
        with connection.cursor() as cursor:
            sql = "insert into accountmanage_account (account_uid,account_YY,account_card,account_belong_id,account_remarks,account_pwd) values ({},{},'{}',{},'{}','{}')".format(
                request.POST.get("account_uid"), request.POST.get("account_YY"), request.POST.get("account_card"),
                request.POST.get("account_belong"), request.POST.get("account_remarks"),
                request.POST.get("account_pwd"))
            print(sql)
            cursor.execute(sql)
            sql_data = cursor.fetchall()  # 获取一条数据, 使用fetchone()
            # colums = [col[0] for col in cursor.description]
            # sql_data = [dict(zip(colums, row)) for row in sql_data]
        print(sql_data)
        result = {"status": True}
        return JsonResponse(result)
    result = {"status": False, "error": form.errors}
    return JsonResponse(result)


@csrf_exempt
def account_edit(request):
    """测试账号详情"""
    aid = request.GET.get("aid")
    row_object = models.Account.objects.filter(id=aid).first()
    print(aid)
    print(row_object)
    if not row_object:
        result = {"status": False, "tips": "编辑失败"}
        return JsonResponse(result)

    form = Account(data=request.POST, instance=row_object)
    print(request.POST)
    if form.is_valid():
        with connection.cursor() as cursor:
            sql = "update accountmanage_account set account_uid='{}',account_YY='{}',account_card='{}',account_belong_id={},account_remarks='{}',account_pwd='{}' where id = {}".format(
                request.POST.get("account_uid"), request.POST.get("account_YY"), request.POST.get("account_card"),
                request.POST.get("account_belong"), request.POST.get("account_remarks"),
                request.POST.get("account_pwd"), aid)
            print(sql)
            cursor.execute(sql)
            sql_data = cursor.fetchall()  # 获取一条数据, 使用fetchone()
            # colums = [col[0] for col in cursor.description]
            # sql_data = [dict(zip(colums, row)) for row in sql_data]
        print(sql_data)
        result = {"status": True}
        return JsonResponse(result)

    result = {"status": False, "error": form.errors}
    return JsonResponse(result)


def account_detail(request):
    """测试账号详情"""
    aid = request.GET.get('aid')
    # print(aid)
    data_dict = models.Account.objects.filter(id=aid).values("account_YY", "account_card", "account_uid",
                                                             "account_belong_id", "account_pwd",
                                                             "account_remarks").first()
    if not data_dict:
        result = {"status": False, "error": "未知错误"}
        return JsonResponse(result)
    print(data_dict)
    result = {"status": True, "data": data_dict}
    # print(data_dict)
    return JsonResponse(result)


def account_operation(request):
    return render(request, "operation_list.html")


def account_option(request):
    with connection.cursor() as cursor:
        sql = "select id,tester_name from accountmanage_testerinfo"
        cursor.execute(sql)
        sql_data = cursor.fetchall()  # 获取一条数据, 使用fetchone()
        # colums = [col[0] for col in cursor.description]
        # sql_data = [dict(zip(colums, row)) for row in sql_data]
    print("sql_data:{}".format(sql_data))
    result = {"status": True, "data": sql_data}
    return JsonResponse(result)


@csrf_exempt
def cal(request):
    if request.method == "GET":
        return render(request, "ssss.html")

    x = int(request.POST.get("x"))
    y = int(request.POST.get("y"))
    op = request.POST.get("op")
    print(x, y, op)
    if op == "add":
        number = x + y
    elif op == "sub":
        number = x - y
    elif op == "mul":
        number = x * y
    elif op == "div":
        number = x / y
    result_dict = {
        "status": True,
        "data": number,
        "x": x,
        "y": y,
        "op": op
    }
    return render(request, "ssss.html", result_dict)


@csrf_exempt
def index(request):
    if request.method == "GET":
        return render(request, "index.html")


@csrf_exempt
def download_excel(request):
    """ 下载模板 """
    file_path = r"AccountManage/static/excel/demo.xlsx"
    f = open(file_path, "rb")
    r = FileResponse(f, as_attachment=True, filename="demo.xlsx")
    return r


@csrf_exempt
def download_schedule(request):
    """ 下载分支测试 """
    file_path = r"AccountManage/static/excel/schedule.xlsx"
    f = open(file_path, "rb")
    r = FileResponse(f, as_attachment=True, filename="schedule.xlsx")
    return r
